"""Build the Toshiba VRF selection BOQ workbook.

Usage:
    python build_boq.py input.json /mnt/user-data/outputs/<Project>_VRF_BOQ.xlsx \
        [discount] [price_list.xlsx]

Diversity has been REMOVED per project instruction: each system's outdoor load
is the plain sum of its indoor required capacities.

Discount lives in cell E2 and is LIVE: every Net formula references $E$2, so
changing E2 in Excel recalculates the whole sheet.

PRICING (privacy-preserving): prices are NEVER stored in the skill. They are
read at runtime from an external price list workbook the user supplies (4th
arg), matched by Toshiba model number, and written into a HIDDEN "Prices" tab
(one row per distinct model used). BOQ Unit Price cells VLOOKUP that tab; Net =
Unit Price * Qty * (1 - $E$2). Delete the Prices tab before sharing to redact.
If no price list is supplied, price cells are left blank.
"""

import json
import sys
import math
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from vrf_data import DEFAULT_DISCOUNT, ACCESSORIES
from engine import (normalize_type, select_idu, select_odu,
                    CASSETTE_TYPES, HI_WALL_TYPE, HIGH_STATIC_TYPE)
from container_calc import calc_containers
from container_data import (CONTAINERS, DEFAULT_CONTAINER, STOWAGE_FACTOR,
                            DIMS, ODU_DIMS)

NAVY = "1F3864"; LIGHT = "D9E1F2"; GREY = "F2F2F2"; ACC = "FCE4D6"; SUM = "E2EFDA"
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10); NORMAL = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _acc_model(name_contains):
    for a in ACCESSORIES:
        if name_contains.lower() in a["name"].lower():
            return a
    return {"model": "Model Not Found", "desc": ""}


def load_price_list(path):
    """Read external price list -> {model: unit_price}. In memory only."""
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    prices = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr_idx = None
        for i, row in enumerate(rows):
            if any(isinstance(c, str) and c.strip() for c in row):
                hdr_idx = i; break
        if hdr_idx is None:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[hdr_idx]]
        model_col = price_col = None
        for j, h in enumerate(header):
            if model_col is None and ("model" in h or "part" in h or "code" in h):
                model_col = j
            if price_col is None and ("price" in h or "cost" in h or "rate" in h):
                price_col = j
        for row in rows[hdr_idx + 1:]:
            if model_col is not None and price_col is not None:
                m = row[model_col] if model_col < len(row) else None
                p = row[price_col] if price_col < len(row) else None
            else:
                m = next((c for c in row if isinstance(c, str) and c.strip()), None)
                p = next((c for c in row if isinstance(c, (int, float))), None)
            if m and isinstance(p, (int, float)):
                prices[str(m).strip()] = float(p)
    return prices


def build(inp, out_path, discount=None, price_list_path=None):
    project = inp.get("project", "VRF Project")
    discount = float(discount if discount is not None
                     else inp.get("discount", DEFAULT_DISCOUNT))
    rows = inp.get("rows", [])
    flags = []

    price_map = load_price_list(price_list_path)
    have_prices = bool(price_map)
    used_models = OrderedDict()  # single-module/IDU/accessory model -> {"qty", "price"}

    # Build a lookup of ODU model -> list of single-module component models.
    # Single-module ODUs (blank modules field) map to themselves.
    import re
    from vrf_data import ODU as _ODU
    _odu_components = {}
    for _m in _ODU:
        mods = (_m.get("modules") or "").strip()
        if mods:
            comps = re.findall(r'M{1,3}Y?-?(MAP\d+HT8P-ME)', mods)
            comps = [c if c.startswith("MAP") else c for c in comps]
            # normalize: catalogue single models are like 'MAP1206HT8P-ME'
            _odu_components[_m["model"]] = comps
        else:
            _odu_components[_m["model"]] = [_m["model"]]

    def odu_modules(model):
        """Return list of single-module component models for an ODU model."""
        return _odu_components.get(model, [model])

    systems = OrderedDict()
    for r0 in rows:
        sysid = str(r0.get("system") or "S1").strip()
        systems.setdefault(sysid, []).append(r0)

    THERM = _acc_model("Thermostat"); PANEL = _acc_model("Cassette panel")
    DRAIN = _acc_model("Drain pump"); BFILT = _acc_model("Back filter")
    n_therm = n_panel = n_drain = n_bfilt = 0
    n_odu = 0
    total_idu = 0; total_idu_kw = 0.0; total_odu_kw_t3 = 0.0

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "VRF BOQ"
    HEAD = ["Unit Tag", "Room", "Type", "Required kW", "Qty",
            "Selected Model", "Rated kW (ea)", "Total kW", "HP",
            "Unit Price", "Net (after disc.)"]
    NCOL = len(HEAD); PRICE_COL = 10; NET_COL = 11

    def register_model(model, qty=1):
        if model and model != "Model Not Found":
            if model not in used_models:
                used_models[model] = {"qty": 0, "price": price_map.get(model)}
            used_models[model]["qty"] += int(qty)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    c = ws.cell(1, 1, f"Toshiba VRF Selection BOQ  \u2014  {project}")
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.cell(2, 4, "Discount Factor").font = BOLD
    dc = ws.cell(2, 5, discount); dc.number_format = "0%"; dc.font = BOLD
    dc.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(2, 7, "Outdoor selection basis").font = BOLD
    ws.cell(2, 8, "T3 @ 46\u00b0C").font = NORMAL
    ws.cell(2, 1, "Change E2 to re-price live").font = Font(italic=True, size=9, color="888888")

    r = 4

    def header_row():
        nonlocal r
        for j, h in enumerate(HEAD, start=1):
            cell = ws.cell(r, j, h)
            cell.font = HDR_FONT; cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = CENTER; cell.border = BORDER
        r += 1
    header_row()

    line_rows = []
    equip_rows = []  # indoor + outdoor unit rows only (for branch-joint base)

    def price_formulas(row_idx, model, qty_col_letter="E", components=None):
        if not model or model == "Model Not Found":
            return False
        if components and len(components) > 1:
            # Unit price = sum of each component module's price from Prices tab
            parts = "+".join(
                f'IFERROR(VLOOKUP("{c}",Prices!$A:$C,3,FALSE),0)' for c in components)
            up = ws.cell(row_idx, PRICE_COL, f'=IF(({parts})=0,"",{parts})')
        else:
            up = ws.cell(row_idx, PRICE_COL,
                         f'=IFERROR(VLOOKUP(F{row_idx},Prices!$A:$C,3,FALSE),"")')
        up.number_format = '#,##0.00'; up.font = NORMAL; up.border = BORDER; up.alignment = CENTER
        net = ws.cell(row_idx, NET_COL,
                      f'=IF(N(J{row_idx})=0,0,J{row_idx}*{qty_col_letter}{row_idx}*(1-$E$2))')
        net.number_format = '#,##0.00'; net.font = NORMAL; net.border = BORDER; net.alignment = CENTER
        return True

    sys_index = 0
    for sysid, srows in systems.items():
        sys_index += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        b = ws.cell(r, 1, f"SYSTEM {sys_index}  ({sysid})")
        b.font = Font(bold=True, color="FFFFFF", size=10)
        b.fill = PatternFill("solid", fgColor="2F5496"); b.alignment = LEFT
        r += 1
        sys_required_t3 = 0.0; iu_counter = 0

        for src in srows:
            ctype, matched = normalize_type(src.get("type"))
            if not matched and src.get("type"):
                flags.append(f"{sysid}: unknown type '{src.get('type')}' -> {ctype}")
            req = float(src.get("required_kw") or 0); qty = int(src.get("qty") or 1)
            selected = select_idu(ctype, req)
            for sel in selected:
                iu_counter += 1
                tag = src.get("tag") or f"IU-{iu_counter:02d}"
                if len(selected) > 1:
                    tag = f"{tag}-{iu_counter}"
                room = src.get("room", "")
                line_kw = sel["cap_kw"] * qty
                total_idu += qty; total_idu_kw += line_kw
                sys_required_t3 += sel["req_kw"] * qty
                vals = [tag, room, ctype, round(sel["req_kw"], 2), qty,
                        sel["model"], sel["cap_kw"], round(line_kw, 2),
                        sel["hp"], "", ""]
                for j, v in enumerate(vals, start=1):
                    cell = ws.cell(r, j, v); cell.font = NORMAL; cell.border = BORDER
                    cell.alignment = LEFT if j in (1, 2, 3, 6) else CENTER
                ws.cell(r, 1).fill = PatternFill("solid", fgColor=GREY)
                register_model(sel["model"], qty)
                if price_formulas(r, sel["model"]):
                    line_rows.append(r); equip_rows.append(r)
                r += 1
                if ctype != HI_WALL_TYPE:
                    n_therm += qty
                if ctype in CASSETTE_TYPES:
                    n_panel += qty
                if ctype == HIGH_STATIC_TYPE:
                    n_drain += qty; n_bfilt += qty

        odus = select_odu(sys_required_t3)  # NO diversity
        for k, od in enumerate(odus, start=1):
            total_odu_kw_t3 += od["t3"]
            n_odu += 1
            otag = f"OU-System {sys_index}" + (f" ({k}/{len(odus)})" if len(odus) > 1 else "")
            vals = [otag, "", od["series"], round(od["req_kw"], 2), 1,
                    od["model"], od["t3"], round(od["t3"], 2), od["hp"], "", ""]
            for j, v in enumerate(vals, start=1):
                cell = ws.cell(r, j, v); cell.font = BOLD; cell.border = BORDER
                cell.alignment = LEFT if j in (1, 2, 3, 6) else CENTER
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            comps = odu_modules(od["model"])
            for c in comps:
                register_model(c, 1)  # price single modules only
            if price_formulas(r, od["model"], components=comps):
                line_rows.append(r); equip_rows.append(r)
            r += 1
        r += 1

    # ACCESSORIES
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    a = ws.cell(r, 1, "ACCESSORIES"); a.font = Font(bold=True, color="FFFFFF", size=11)
    a.fill = PatternFill("solid", fgColor="C55A11"); a.alignment = LEFT
    r += 1
    acc_hdr = ["Item", "", "", "", "Qty", "Model", "", "", "", "Unit Price", "Net (after disc.)"]
    for j, h in enumerate(acc_hdr, start=1):
        cell = ws.cell(r, j, h); cell.font = BOLD; cell.border = BORDER
        cell.alignment = CENTER; cell.fill = PatternFill("solid", fgColor=ACC)
    r += 1

    def acc_line(item, qty, model):
        nonlocal r
        ws.cell(r, 1, item).font = NORMAL
        ws.cell(r, 5, qty).font = NORMAL
        ws.cell(r, 6, model).font = NORMAL
        for j in range(1, NCOL + 1):
            ws.cell(r, j).border = BORDER
            ws.cell(r, j).alignment = LEFT if j in (1, 6) else CENTER
        if isinstance(qty, int) and qty > 0:
            register_model(model, qty)
            if price_formulas(r, model):
                line_rows.append(r)
        r += 1

    acc_line("Thermostat (all IDU except Hi-Wall)", n_therm, THERM["model"])
    acc_line("Cassette Panel (per cassette IDU)", n_panel, PANEL["model"])
    acc_line("Drain Pump (per High Static Ducted)", n_drain, DRAIN["model"])
    acc_line("Back Filter (per High Static Ducted)", n_bfilt, BFILT["model"])
    n_bms = math.ceil(total_idu / 128) if total_idu else 0
    acc_line("BMS Interface (1 per 128 indoor units)", n_bms, "BMS-IFBN1281U-E")

    ws.cell(r, 1, "Branch Joints (Refnet)").font = NORMAL
    ws.cell(r, 6, "\u2014").font = NORMAL
    branch_net_row = None
    if equip_rows:
        net_refs = "+".join(f"K{i}" for i in equip_rows)
        bj = ws.cell(r, NET_COL, f"=0.05*({net_refs})")
        bj.number_format = '#,##0.00'; bj.font = NORMAL
        branch_net_row = r
    else:
        ws.cell(r, NET_COL, "5% of IDU+ODU value").font = NORMAL
    for j in range(1, NCOL + 1):
        ws.cell(r, j).border = BORDER
        ws.cell(r, j).alignment = LEFT if j in (1, 6) else CENTER
    r += 2

    # ===================== GRAND TOTALS (separate, above summary) =====================
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    g = ws.cell(r, 1, "GRAND TOTALS"); g.font = Font(bold=True, color="FFFFFF", size=12)
    g.fill = PatternFill("solid", fgColor="1F3864"); g.alignment = LEFT
    ws.row_dimensions[r].height = 20
    r += 1

    gt_nodisc_row = gt_disc_row = None
    GT_VAL_COL = NET_COL  # 11 (K) — the wide Net column, fits large numbers

    def gt_line(label, value, fmt='#,##0.00', fill="DDEBF7", style="normal"):
        """style: 'big' (large white on fill), 'muted' (small grey, no fill), 'normal'."""
        nonlocal r
        if style == "big":
            lf = Font(bold=True, size=13, color="FFFFFF")
            ws.row_dimensions[r].height = 24
        elif style == "muted":
            lf = Font(size=9, color="808080")
            ws.row_dimensions[r].height = 16
        else:
            lf = Font(bold=True, size=10)
        # label across columns A:I, value in J:K (merged, wide & right-aligned)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=GT_VAL_COL - 2)
        lc = ws.cell(r, 1, label); lc.font = lf
        lc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=GT_VAL_COL - 1, end_row=r, end_column=GT_VAL_COL)
        vc = ws.cell(r, GT_VAL_COL - 1, value); vc.font = lf
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            vc.number_format = fmt
        for j in range(1, NCOL + 1):
            ws.cell(r, j).border = BORDER
            if style != "muted":
                ws.cell(r, j).fill = PatternFill("solid", fgColor=fill)
        this = r
        r += 1
        return this

    if line_rows:
        # Net references already embed (1 - E2). Gross = Net / (1 - E2).
        all_nets = "+".join(f"K{i}" for i in line_rows)
        if branch_net_row:
            all_nets += f"+K{branch_net_row}"
        net_expr = f"({all_nets})"
        gt_disc_row = gt_line("GRAND TOTAL  (with discount)",
                              f"={net_expr}", fill="538135", style="big")
        gt_nodisc_row = gt_line("Grand Total (without discount)",
                                f"=IF((1-$E$2)=0,0,{net_expr}/(1-$E$2))",
                                style="muted")
        gt_line("Discount applied", "=E2", fmt="0%", fill="FFF2CC")
    else:
        gt_line("Grand Total", "Fill Unit Price in the 'Prices' tab", fmt=None, fill="F2F2F2")
    r += 1  # spacer

    # ===================== SUMMARY =====================
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    s = ws.cell(r, 1, "SUMMARY"); s.font = Font(bold=True, color="FFFFFF", size=12)
    s.fill = PatternFill("solid", fgColor="538135"); s.alignment = LEFT
    ws.row_dimensions[r].height = 20
    r += 1

    def sum_line(label, value, fmt=None, bold=False, fill=SUM):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NET_COL - 2)
        lc = ws.cell(r, 1, label); lc.font = BOLD if bold else NORMAL
        lc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=NET_COL - 1, end_row=r, end_column=NET_COL)
        cell = ws.cell(r, NET_COL - 1, value); cell.font = BOLD if bold else NORMAL
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            cell.number_format = fmt
        for j in range(1, NCOL + 1):
            ws.cell(r, j).border = BORDER
            ws.cell(r, j).fill = PatternFill("solid", fgColor=fill)
        this = r
        r += 1
        return this

    VAL_COL_LETTER = get_column_letter(NET_COL - 1)  # 'J'
    TR = 3.517
    n_systems = len(systems)
    sum_line("No. of Systems", n_systems, bold=True)
    sum_line("Total Indoor Units (qty)", total_idu, bold=True)
    sum_line("Total Outdoor Units (qty)", n_odu, bold=True)
    sum_line("Total Indoor Capacity (kW, rated)", round(total_idu_kw, 1))
    sum_line("Total Indoor Capacity (TR)", round(total_idu_kw / TR, 1))
    sum_line("Total Outdoor Capacity (kW, T3)", round(total_odu_kw_t3, 1))
    out_tr = total_odu_kw_t3 / TR if total_odu_kw_t3 else 0
    out_tr_row = sum_line("Total Outdoor Capacity (TR, T3)", round(out_tr, 1))
    sum_line("Discount Factor", "=E2", "0%")

    # Price per TR (bold, color-coded). Based on WITH-discount net grand total
    # over outdoor TR. <900 green, 1000-1200 yellow, >=1200 red.
    if line_rows and gt_disc_row and out_tr:
        vL = VAL_COL_LETTER
        ppt_row = sum_line("Price per TR",
                           f"=IF({vL}{out_tr_row}=0,0,{vL}{gt_disc_row}/{vL}{out_tr_row})",
                           fmt='#,##0.00', bold=True, fill="FFFFFF")
        ws.cell(ppt_row, NET_COL - 1).font = Font(bold=True, size=11)
        cell_ref = f"{vL}{ppt_row}"
        from openpyxl.formatting.rule import CellIsRule
        green = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        red = PatternFill("solid", fgColor="FFC7CE")
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator="lessThan", formula=["900"], fill=green))
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator="greaterThanOrEqual", formula=["1200"], fill=red))
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator="between", formula=["1000", "1200"], fill=yellow))
    else:
        sum_line("Price per TR", "Fill prices to compute", bold=True, fill="FFFFFF")

    if line_rows:
        sum_line("Pricing source", "Fill Unit Price in the 'Prices' tab")
    else:
        sum_line("Pricing", "No models to price")

    widths = [16, 16, 24, 11, 6, 22, 12, 10, 7, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ---------------- PRICES TAB (your input surface) ----------------
    missing = []
    if used_models:
        pw = wb.create_sheet("Prices")
        ph = ["Model", "Qty", "Unit Price", "Line Total"]
        for j, h in enumerate(ph, start=1):
            hc = pw.cell(1, j, h); hc.font = HDR_FONT
            hc.fill = PatternFill("solid", fgColor=NAVY); hc.alignment = CENTER; hc.border = BORDER
        pr = 2
        for model, info in used_models.items():
            pw.cell(pr, 1, model).border = BORDER
            qy = pw.cell(pr, 2, info["qty"]); qy.alignment = CENTER; qy.border = BORDER
            up = pw.cell(pr, 3)
            if info["price"] is not None:
                up.value = info["price"]
            else:
                up.fill = PatternFill("solid", fgColor="FFF2CC")  # highlight: fill me
                if have_prices:
                    missing.append(model)
            up.number_format = '#,##0.00'; up.alignment = CENTER; up.border = BORDER
            lt = pw.cell(pr, 4, f'=IF(N(C{pr})=0,"",B{pr}*C{pr})')
            lt.number_format = '#,##0.00'; lt.alignment = CENTER; lt.border = BORDER
            pr += 1
        last_item = pr - 1
        # Equipment subtotal (gross, before branch joints & discount)
        pw.cell(pr, 1, "Equipment Subtotal (gross)").font = BOLD
        sub = pw.cell(pr, 4, f'=SUM(D2:D{last_item})')
        sub.font = BOLD; sub.number_format = '#,##0.00'
        for j in range(1, 5):
            pw.cell(pr, j).border = BORDER
            pw.cell(pr, j).fill = PatternFill("solid", fgColor="F2F2F2")
        sub_row = pr; pr += 1
        # Branch joints = 5% of indoor+outdoor equipment (gross), excluding
        # accessories. Mirrors the BOQ equipment net, grossed up by the discount.
        pw.cell(pr, 1, "Branch Joints (5% of IDU+ODU)").font = NORMAL
        if equip_rows:
            equip_net = "+".join(f"'VRF BOQ'!K{i}" for i in equip_rows)
            bj = pw.cell(pr, 4,
                         f"=IF((1-'VRF BOQ'!$E$2)=0,0,0.05*({equip_net})/(1-'VRF BOQ'!$E$2))")
        else:
            bj = pw.cell(pr, 4, 0)
        bj.font = NORMAL; bj.number_format = '#,##0.00'
        for j in range(1, 5):
            pw.cell(pr, j).border = BORDER
            pw.cell(pr, j).fill = PatternFill("solid", fgColor="F2F2F2")
        bj_row = pr; pr += 1
        # Project total (gross) = equipment + branch joints
        pw.cell(pr, 1, "PROJECT TOTAL (gross)").font = BOLD
        gross = pw.cell(pr, 4, f'=D{sub_row}+D{bj_row}')
        gross.font = BOLD; gross.number_format = '#,##0.00'
        for j in range(1, 5):
            pw.cell(pr, j).border = BORDER
            pw.cell(pr, j).fill = PatternFill("solid", fgColor="FCE4D6")
        gross_row = pr; pr += 1
        # GRAND TOTAL WITHOUT discount (= gross) — smaller, grey, not highlighted
        wod = pw.cell(pr, 1, "Grand Total (without discount)")
        wod.font = Font(size=9, color="808080"); wod.alignment = LEFT
        wodv = pw.cell(pr, 4, f'=D{gross_row}')
        wodv.font = Font(size=9, color="808080"); wodv.number_format = '#,##0.00'
        for j in range(1, 5):
            pw.cell(pr, j).border = BORDER
        pr += 1
        # GRAND TOTAL WITH discount — big & highlighted (uses BOQ discount cell)
        wd = pw.cell(pr, 1, "GRAND TOTAL (with discount)")
        wd.font = Font(bold=True, size=13, color="FFFFFF"); wd.alignment = LEFT
        wdv = pw.cell(pr, 4, f"=D{gross_row}*(1-'VRF BOQ'!$E$2)")
        wdv.font = Font(bold=True, size=13, color="FFFFFF"); wdv.number_format = '#,##0.00'
        for j in range(1, 5):
            pw.cell(pr, j).border = BORDER
            pw.cell(pr, j).fill = PatternFill("solid", fgColor="538135")
        pw.row_dimensions[pr].height = 22
        pr += 1
        pw.column_dimensions["A"].width = 30
        pw.column_dimensions["B"].width = 8
        pw.column_dimensions["C"].width = 14
        pw.column_dimensions["D"].width = 16
        pw.sheet_state = "visible"

    # ---------------- CONTAINER LOADING TAB ----------------
    # ODU single modules cannot be stacked (floor-lane driven); IDU can be
    # stacked (volume driven). Containers = max(ODU floor, IDU volume, weight).
    # used_models already holds single MAP* module qtys + IDU qtys; accessories
    # have no shipping dims and are skipped automatically.
    container_summary = None
    ship_qtys = {m: info["qty"] for m, info in used_models.items() if m in DIMS}
    if ship_qtys:
        container_summary = calc_containers(ship_qtys, DEFAULT_CONTAINER,
                                            STOWAGE_FACTOR)
        cw = wb.create_sheet("Container Loading")
        # Title
        cw.merge_cells("A1:G1")
        t = cw.cell(1, 1, f"Container Loading  \u2014  {project}")
        t.font = Font(bold=True, color="FFFFFF", size=14)
        t.fill = PatternFill("solid", fgColor=NAVY)
        t.alignment = Alignment(horizontal="center", vertical="center")
        cw.row_dimensions[1].height = 26
        cw.merge_cells("A2:G2")
        cw.cell(2, 1, "ODU modules CANNOT be stacked (floor-lane driven) \u00b7 "
                      "IDU CAN be stacked (volume driven) \u00b7 weight always "
                      "checked. Containers = max of the three drivers.").font = \
            Font(italic=True, size=9, color="808080")

        # Reference block (editable yellow inputs)
        rr = 4
        cw.cell(rr, 1, "Container type").font = BOLD
        ct = cw.cell(rr, 3, DEFAULT_CONTAINER); ct.font = BOLD
        ct.fill = PatternFill("solid", fgColor="FFF2CC"); ct.alignment = CENTER
        rr += 1
        cw.cell(rr, 1, "Stowage / pack factor (IDU)").font = BOLD
        sf = cw.cell(rr, 3, STOWAGE_FACTOR); sf.number_format = "0%"; sf.font = BOLD
        sf.fill = PatternFill("solid", fgColor="FFF2CC"); sf.alignment = CENTER
        rr += 1
        cont = CONTAINERS[DEFAULT_CONTAINER]
        cw.cell(rr, 1, "Interior L \u00d7 W \u00d7 H (mm)").font = NORMAL
        cw.cell(rr, 3, f"{cont['len']} \u00d7 {cont['wid']} \u00d7 {cont['hgt']}").alignment = CENTER
        rr += 1
        cw.cell(rr, 1, "Usable volume / container (m\u00b3)").font = NORMAL
        cw.cell(rr, 3, container_summary["usable_vol_per_container_m3"]).alignment = CENTER
        rr += 1
        cw.cell(rr, 1, "Max payload / container (kg)").font = NORMAL
        cw.cell(rr, 3, container_summary["payload_per_container_kg"]).alignment = CENTER
        rr += 2

        # Manifest table
        mh = ["Group", "Model", "Qty", "Stack?", "Unit Vol (m\u00b3)",
              "Unit Wt (kg)", "Total Wt (kg)"]
        for j, h in enumerate(mh, start=1):
            hc = cw.cell(rr, j, h); hc.font = HDR_FONT
            hc.fill = PatternFill("solid", fgColor=NAVY); hc.alignment = CENTER; hc.border = BORDER
        rr += 1
        for model, qty in ship_qtys.items():
            d = DIMS[model]
            is_odu = model in ODU_DIMS
            vals = ["ODU module" if is_odu else "Indoor", model, qty,
                    "No" if is_odu else "Yes", d["vol"], d["wt"], d["wt"] * qty]
            for j, v in enumerate(vals, start=1):
                c = cw.cell(rr, j, v); c.font = NORMAL; c.border = BORDER
                c.alignment = LEFT if j in (1, 2) else CENTER
                if is_odu:
                    c.fill = PatternFill("solid", fgColor=LIGHT)
            rr += 1
        rr += 1

        # Drivers + result
        cw.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
        d0 = cw.cell(rr, 1, "CONTAINER DRIVERS"); d0.font = Font(bold=True, color="FFFFFF", size=11)
        d0.fill = PatternFill("solid", fgColor="2F5496"); d0.alignment = LEFT
        rr += 1

        def drv(label, val, note="", governing=False):
            nonlocal rr
            lc = cw.cell(rr, 1, label); lc.font = BOLD if governing else NORMAL
            cw.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
            vc = cw.cell(rr, 5, val); vc.alignment = CENTER
            vc.font = BOLD if governing else NORMAL; vc.number_format = "0.000"
            cw.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=7)
            nc = cw.cell(rr, 6, note); nc.font = Font(size=9, color="808080"); nc.alignment = LEFT
            fill = "C6EFCE" if governing else "F2F2F2"
            for j in range(1, 8):
                cw.cell(rr, j).border = BORDER
                cw.cell(rr, j).fill = PatternFill("solid", fgColor=fill)
            rr += 1

        g = container_summary["governing_driver"]
        drv("Containers by ODU floor lanes",
            container_summary["containers_by_odu_floor"],
            f"{container_summary['odu_modules']} ODU modules, non-stackable",
            g == "odu_floor")
        drv("Containers by IDU volume",
            container_summary["containers_by_idu_volume"],
            f"{container_summary['idu_units']} indoor units, stacked",
            g == "idu_volume")
        drv("Containers by weight",
            container_summary["containers_by_weight"],
            f"{container_summary['total_weight_kg']:.0f} kg total",
            g == "weight")
        rr += 1
        cw.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
        rl = cw.cell(rr, 1, f"CONTAINERS REQUIRED ({DEFAULT_CONTAINER})")
        rl.font = Font(bold=True, color="FFFFFF", size=13); rl.alignment = LEFT
        cw.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=7)
        rv = cw.cell(rr, 5, container_summary["containers_required"])
        rv.font = Font(bold=True, color="FFFFFF", size=13); rv.alignment = CENTER
        for j in range(1, 8):
            cw.cell(rr, j).border = BORDER
            cw.cell(rr, j).fill = PatternFill("solid", fgColor="538135")
        cw.row_dimensions[rr].height = 24

        for col, w in zip("ABCDEFG", [16, 22, 8, 8, 12, 12, 14]):
            cw.column_dimensions[col].width = w
        cw.sheet_state = "visible"

    wb.save(out_path)
    return {
        "project": project, "systems": len(systems),
        "total_indoor_units": total_idu,
        "total_indoor_kw": round(total_idu_kw, 1),
        "total_outdoor_kw_t3": round(total_odu_kw_t3, 1),
        "thermostats": n_therm, "cassette_panels": n_panel,
        "drain_pumps": n_drain, "back_filters": n_bfilt,
        "discount": discount, "priced": have_prices,
        "total_outdoor_units": n_odu, "n_systems": len(systems),
        "models_used": len(used_models),
        "models_missing_price": missing,
        "container_loading": container_summary,
        "flags": flags, "output": out_path,
    }


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: python build_boq.py input.json output.xlsx [discount] [price_list.xlsx]")
        sys.exit(1)
    inp = json.load(open(sys.argv[1]))
    disc = float(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[3] not in ("", "-") else None
    plist = sys.argv[4] if len(sys.argv) > 4 else None
    res = build(inp, sys.argv[2], disc, plist)
    print(json.dumps(res, indent=2))
